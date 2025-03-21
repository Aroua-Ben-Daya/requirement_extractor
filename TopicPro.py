import pdfplumber
import pandas as pd
import re
import os
from openpyxl import load_workbook

# File paths
#D:/UNIVERSITE D'AIX MARSEILLE/Day 1 03_03_2025/X2R3/X2R3-T4_3-D-SMD-008-19_-_D4.2Part3-SystemSpecification.pdf
#D:/UNIVERSITE D'AIX MARSEILLE/Day 1 03_03_2025/X2R5/X2R5-T4_2-D-SMD-003-23_-_D41Part3SystemSpecification.pdf
pdf_path = "D:/UNIVERSITE D'AIX MARSEILLE/Day 1 03_03_2025/X2R5/X2R5-T4_2-D-SMD-003-23_-_D41Part3SystemSpecification.pdf"
excel_template = "D:/UNIVERSITE D'AIX MARSEILLE/Day 1 03_03_2025/req Eng.xlsx"

# Generate dynamic output file name based on input PDF
pdf_filename = os.path.basename(pdf_path).replace(".pdf", "")
output_excel = f"C:/Users/aroua/Downloads/{pdf_filename}_rResult.xlsx"

# Load Excel template to preserve formatting
wb = load_workbook(excel_template)
ws = wb.active

# Extract column headers from the template
columns_needed = [cell.value for cell in ws[1] if cell.value]
required_columns = ["Topic", "Requirement ID", "Description", "Traceability"]

# Function to extract requirements from PDF
def extract_requirements(pdf_path):
    requirements = []
    req_pattern = re.compile(r"(REQ-[A-Za-z0-9]+-\d+) \[(.*?)\]")  # Matches requirement IDs and traceability

    topic_pattern = re.compile(r"^(?:\d+\.\d+\.?\d*\s*)?([A-Za-z][A-Za-z0-9 \-]+)$")  


    with pdfplumber.open(pdf_path) as pdf:
        current_topic = "Unknown"
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(layout=True)  # Preserve layout
            if text:
                lines = text.split("\n")
                for i, line in enumerate(lines):
                    topic_match = topic_pattern.match(line)
                    if topic_match:
                        current_topic = topic_match.group(1).strip()
                    
                    req_match = req_pattern.search(line)
                    if req_match:
                        req_id = req_match.group(1)
                        traceability = f"[{req_match.group(2)}]"  # Ensure traceability is formatted correctly

                        description = extract_description(lines, i)
                        '''if description:
                            print(f"✅ Found requirement on page {page_num}: {req_id}")
                            requirements.append((current_topic, req_id, description, traceability))'''
                        if description:
                            print(f"✅ Page {page_num}: {req_id} | Topic: {topic} | Traceability: {traceability}")
                            requirements.append((topic, req_id, description, traceability))
                            
    return requirements

# Extract description until 'Rationale' section
def extract_description(lines, start_idx):
    description = []
    for i in range(start_idx + 1, len(lines)):
        if "Rationale:" in lines[i]:
            break
        description.append(lines[i].strip())
    return " ".join(description).strip()

# Extract requirements
data = extract_requirements(pdf_path)

# Convert to DataFrame
extracted_df = pd.DataFrame(data, columns=required_columns)

# Ensure all required columns exist in the DataFrame
for col in columns_needed:
    if col not in extracted_df.columns:
        extracted_df[col] = ""  # Fill missing columns with empty values

# Write extracted data into the formatted Excel template
for row_idx, row in extracted_df.iterrows():
    for col_idx, col_name in enumerate(required_columns):
        ws[f"{chr(65 + col_idx)}{row_idx + 2}"] = row[col_name]

# Save to new Excel file while preserving formatting
wb.save(output_excel)

print(f"✅ Extracted requirements saved to {output_excel} with original structure and formatting!")
