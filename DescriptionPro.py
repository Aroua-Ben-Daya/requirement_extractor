import pdfplumber
import pandas as pd
import re
import os
from openpyxl import load_workbook

# File paths
pdf_path = "D:/UNIVERSITE D'AIX MARSEILLE/Day 1 03_03_2025/X2R5/X2R5-T4_2-D-SMD-003-23_-_D41Part3SystemSpecification.pdf"
excel_template = "D:/UNIVERSITE D'AIX MARSEILLE/Day 1 03_03_2025/req Eng.xlsx"

# Generate dynamic output file name based on input PDF
pdf_filename = os.path.basename(pdf_path).replace(".pdf", "")
output_excel = f"C:/Users/aroua/Downloads/{pdf_filename}_testResult.xlsx"

# Load Excel template to preserve formatting
wb = load_workbook(excel_template)
ws = wb.active

# Extract column headers from the template
columns_needed = [cell.value for cell in ws[1] if cell.value]
required_columns = ["Topic", "Requirement ID", "Description", "Traceability"]

# Patterns for requirement extraction
#req_pattern = re.compile(r"(REQ-[A-Za-z0-9]+-\d+)\s*(\[[^\]]+\])?")  # Matches requirement IDs and optional traceability
#topic_pattern = re.compile(r"^\d+\.\d+(?:\.\d+)?\s+([A-Za-z][A-Za-z0-9 \-]+)")  # Matches section titles like "1.2.3 Topic"
req_pattern = re.compile(r"(REQ-[A-Za-z0-9]+-\d+)\s*(\[[^\]]+\])?")  # Matches requirement IDs and optional traceability
topic_pattern = re.compile(r"^(?:\d+\.\d+\.?\d*\s*)?([A-Za-z][A-Za-z0-9 \-]+)$")  

def extract_requirements(pdf_path):
    requirements = []  # Ensure the list is cleared for each new PDF
    current_topic = None
    last_valid_topic = "Unknown"  # Default topic if no valid topic is found

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(layout=True)
            if text:
                lines = text.split("\n")
                for i, line in enumerate(lines):
                    # Topic Detection - Ensure accurate assignment
                    topic_match = topic_pattern.match(line)
                    if topic_match:
                        current_topic = topic_match.group(1).strip()
                        last_valid_topic = current_topic  # Update last valid topic
                    
                    # Requirement & Traceability Extraction
                    req_match = req_pattern.search(line)
                    if req_match:
                        req_id = req_match.group(1)
                        traceability = req_match.group(2) if req_match.group(2) else "[Not Provided]"
                        description = extract_description(lines, i)
                        topic = current_topic if current_topic else last_valid_topic

                        if description:
                            print(f"✅ Page {page_num}: {req_id} | Topic: {topic} | Traceability: {traceability}")
                            requirements.append((topic, req_id, description, traceability))
    return requirements

# Extract description until 'Rationale' section
def extract_description(lines, start_idx):
    description = []
    for i in range(start_idx + 1, len(lines)):
        if "Rationale:" in lines[i]:
            break
        description.append(lines[i].strip())
    return " ".join(description).strip()

# Extract requirements
data = extract_requirements(pdf_path)

# Convert to DataFrame
extracted_df = pd.DataFrame(data, columns=required_columns)

# Ensure all required columns exist in the DataFrame
for col in columns_needed:
    if col not in extracted_df.columns:
        extracted_df[col] = ""  # Fill missing columns with empty values

# Clear previous Excel content before writing new data
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(required_columns)):
    for cell in row:
        cell.value = None

# Write extracted data into the formatted Excel template
for row_idx, row in extracted_df.iterrows():
    for col_idx, col_name in enumerate(required_columns):
        ws[f"{chr(65 + col_idx)}{row_idx + 2}"] = row[col_name]

# Save to new Excel file while preserving formatting
wb.save(output_excel)

print(f"✅ Extracted requirements saved to {output_excel} with original structure and formatting!")
