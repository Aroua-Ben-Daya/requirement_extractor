import pdfplumber
import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# File paths
pdf_path = "D:/UNIVERSITE D'AIX MARSEILLE/Day 1 03_03_2025/X2R5/X2R5-T4_2-D-SMD-003-23_-_D41Part3SystemSpecification.pdf"
excel_template = "D:/UNIVERSITE D'AIX MARSEILLE/Day 1 03_03_2025/req Eng.xlsx"
# Generate output file name
pdf_filename = os.path.basename(pdf_path).replace(".pdf", "")
output_excel = f"C:/Users/aroua/Downloads/{pdf_filename}_rResult.xlsx"

# Load Excel template
wb = load_workbook(excel_template)
ws = wb.active

# Extract column headers from template
columns_needed = [cell.value for cell in ws[1] if cell.value]
required_columns = ["Topic", "Requirement ID", "Description", "Traceability"]

# Patterns for extraction
topic_pattern = re.compile(r"^\s*(\d+\.\d+)\s+([A-Za-z][A-Za-z0-9 \-]+)")
req_pattern = re.compile(r"(REQ-[A-Za-z0-9]+-\d+)\s*(\[[^\]]+\])?")
traceability_pattern = re.compile(r"\[([A-Za-z0-9\s\-\.:]+)\]")  # More flexible pattern
footer_pattern = re.compile(r"GA\s\d+\s+Page\s\d+\s+of\s+\d+", re.IGNORECASE)  # Removes footers

# Function to extract requirements
def extract_requirements(pdf_path):
    requirements = []
    current_topic = "Unknown"
    last_traceability = "[Not Provided]"  # Default traceability if none found

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(layout=True)
            if text:
                lines = text.split("\n")

                for idx, line in enumerate(lines):
                    # Remove footers
                    line = footer_pattern.sub("", line).strip()

                    # Detect major section topics (3.1, 3.2, etc.)
                    topic_match = topic_pattern.match(line)
                    if topic_match:
                        section_number = topic_match.group(1).strip()
                        topic_name = topic_match.group(2).strip()
                        current_topic = f"{section_number} - {topic_name}"
                        print(f"✅ Detected Topic: {current_topic}")

                    # Detect traceability anywhere in the text
                    traceability_match = traceability_pattern.search(line)
                    if traceability_match:
                        last_traceability = traceability_match.group(1).strip()
                        print(f"🔵 Found Traceability: {last_traceability}")

                    # Extract requirement ID only from header bars
                    if "Mandatory" in line or "Optional" in line:
                        req_match = req_pattern.search(line)
                        if req_match:
                            req_id = req_match.group(1)
                            traceability = req_match.group(2) if req_match.group(2) else last_traceability
                            traceability = traceability.strip("[]").strip()  # Clean brackets

                            # Extract description
                            description = extract_description(lines, idx + 1)

                            # ✅ Debugging: Ensure traceability is stored correctly
                            print(f"📌 Storing: {req_id} | {current_topic} | {traceability}")

                            # Append to requirements list
                            requirements.append((current_topic, req_id, description, traceability))
    
    return requirements

# Function to extract description
def extract_description(lines, start_idx):
    description = []
    for i in range(start_idx, len(lines)):
        if "Rationale:" in lines[i] or "Guidance:" in lines[i]:  # Stop at Rationale/Guidance
            break
        clean_line = footer_pattern.sub("", lines[i]).strip()  # Remove footer
        if clean_line:
            description.append(clean_line)
    return "\n".join(description).strip()  # Keep structured formatting

# Extract requirements
data = extract_requirements(pdf_path)

# Convert to DataFrame
extracted_df = pd.DataFrame(data, columns=required_columns)

# ✅ Debugging: Check if traceability is being stored correctly
print(extracted_df.head())  # Print first few rows to verify traceability column

# Ensure all required columns exist
for col in columns_needed:
    if col not in extracted_df.columns:
        extracted_df[col] = ""  # Fill missing columns

# Write extracted data to Excel
for row_idx, row in extracted_df.iterrows():
    for col_idx, col_name in enumerate(required_columns):
        ws[f"{chr(65 + col_idx)}{row_idx + 2}"] = row[col_name]

# Formatting: Wrap text and auto-adjust cells
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(required_columns)):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)  # Enable text wrapping

# Adjust row height automatically
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):  # Only Description column
    for cell in row:
        ws.row_dimensions[cell.row].height = None  # Auto-adjust row height

# Adjust column width (Limit to max 50 for better readability)
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter  # Get column letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = min(max_length + 5, 50)  # Limit width to 50 characters

# Save final output
wb.save(output_excel)

print(f"✅ Extraction completed! File saved at {output_excel}")
