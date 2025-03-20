import pdfplumber
import pandas as pd
import re
import os
from openpyxl import load_workbook

# Updated file paths
pdf_path = "C:/Users/aroua/AppData/Local/Temp/11c88991-5b2f-430e-996a-3bf8c01428ef_X2R3-T4_3-D-SMD-033-02_-_D4.2MovingBlockSpecifications.zip.8ef/X2R3-T4_3-D-SMD-008-19_-_D4.2Part3-SystemSpecification.pdf"
excel_template = "D:/UNIVERSITE D'AIX MARSEILLE/Day 1 03_03_2025/req Eng.xlsx"
output_excel = "C:/Users/aroua/Downloads/extracted_requirements.xlsx"

# Load Excel template to preserve formatting
wb = load_workbook(excel_template)
ws = wb.active

# Extract column headers from the template
columns_needed = [cell.value for cell in ws[1] if cell.value]
required_columns = ["Requirement ID", "Description", "Topic", "Selection (pertinent ?)", "Category", "Variant MB", "Comment / Missing Information", "Operational mode"]

# Function to extract requirements from PDF
def extract_requirements(pdf_path):
    requirements = []
    pattern = re.compile(r"(REQ-[A-Za-z0-9]+-\d+)")  # Matches requirement IDs like REQ-TTD-4

    with pdfplumber.open(pdf_path) as pdf:
        current_topic = "Unknown"
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(layout=True)  # Preserve layout for structured text
            if text:
                lines = text.split("\n")
                for line in lines:
                    if re.match(r"^[A-Z][A-Za-z ]+:$", line) and len(line.split()) < 7:  # Detect section headers
                        current_topic = line.replace(":", "").strip()
                    match = pattern.search(line)
                    if match:
                        req_id = match.group(1)
                        req_text = extract_full_requirement(text, req_id)
                        refined_desc = refine_description(req_text)
                        category = categorize_requirement(req_text)
                        if refined_desc:
                            print(f"✅ Found requirement on page {page_num}: {req_id}")
                            requirements.append((current_topic, req_id, refined_desc, "Yes", category, "MB1", "Needs review", "nominal"))
    return requirements

# Extract full requirement text including multi-line descriptions
def extract_full_requirement(text, req_id):
    pattern = rf"({req_id}.*?)(?=\nREQ-[A-Za-z0-9]+-\d+|\Z)"
    match = re.search(pattern, text, re.DOTALL)
    if match:
        description = match.group(1).strip()
        return description.replace(req_id, "", 1).strip()
    return "Description not found"

# Refine requirement descriptions for clarity and readability
def refine_description(description):
    description = re.sub(r"\s+", " ", description).strip()
    description = description.replace("shall", "must")
    description = description.replace("The L3 Trackside", "The system")
    description = re.sub(r"\b(e\.g\.|i\.e\.|etc\.)", "for example", description)
    return description

# Categorize requirements based on keywords
def categorize_requirement(description):
    if "location" in description.lower():
        return "Train Location"
    elif "movement authority" in description.lower():
        return "Movement Authority"
    elif "track status" in description.lower():
        return "Track Status"
    elif "shunting" in description.lower():
        return "Shunting"
    return "General"

# Extract requirements
data = extract_requirements(pdf_path)

# Convert to DataFrame
extracted_df = pd.DataFrame(data, columns=required_columns)

# Ensure all required columns exist in the DataFrame
for col in columns_needed:
    if col not in extracted_df.columns:
        extracted_df[col] = ""

# Write extracted data into the formatted Excel template
for row_idx, row in extracted_df.iterrows():
    for col_idx, col_name in enumerate(required_columns):
        ws[f"{chr(65 + col_idx)}{row_idx + 2}"] = row[col_name]

# Save to new Excel file while preserving formatting
wb.save(output_excel)

print(f"✅ Extracted requirements saved to {output_excel} with original structure and formatting!")
